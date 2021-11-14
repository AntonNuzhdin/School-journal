import sqlite3
from openpyxl import *

put = 'baza_for_sait.db'
put_ex = 'students.xlsx'


wb = load_workbook(filename = put_ex)
xl = wb['Данные']
stroka = 1
man = list()
while xl.cell(stroka, 1).value:
    man.append(xl.cell(stroka, 1).value)
    stroka += 1

con = sqlite3.connect(put)
cur = con.cursor()
result = cur.execute(f'select id, ФИО FROM students').fetchall()
# result = cur.execute(f'DELETE FROM teacher_notes WHERE class = ({zapros})').fetchall()
for element in result:
    if element[1] not in man:
        print(element)

# con.commit()
con.close()
