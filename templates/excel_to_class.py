# import pandas as pd
from openpyxl import *
import sqlite3

put = '/Users/antonnuzhdin/Yandex.Disk.localized/new_abulance/baza_for_sait.db'
# wb = load_workbook('/Users/antonnuzhdin/Downloads/rasp.xlsx')
# sh = wb.sheetnames()
# print(wb["Sheet1"])
# excel_data_df = pd.read_excel('/Users/antonnuzhdin/Downloads/rasp.xlsx', sheet_name='Лист1',  usecols=['9з', 'Абакумов Артём Алексеевич'])
# print('Excel Sheet to Dict:', excel_data_df.to_dict(orient='record'))
# print(excel_data_df)
# file = '/Users/antonnuzhdin/Downloads/rasp.xlsx'
#
# xl = pd.ExcelFile(file)
#
# print(xl.sheet_names)
# from openpyxl import *
# source_file = load_workbook("/Users/antonnuzhdin/Downloads/rasp.xlsx")
# sheet = source_file["B"]
#
# cell_value = sheet.cell(25,52)
# print(cell_value)
#
# source_file.close()
con = sqlite3.connect(put)

# Создание курсора
cur = con.cursor()
#находим по логину пользователя
wb = load_workbook(filename = '/Users/antonnuzhdin/Downloads/rasp.xlsx')
klass_list = wb['Лист1']
stroka = 1
while klass_list.cell(stroka, 2).value:
    fio = klass_list.cell(stroka, 1).value
    cl = klass_list.cell(stroka, 2).value
    rez = cur.execute(f'''INSERT INTO admission(fio, to_class) VALUES (\'{fio}\', \'{cl}\')''')
    con.commit()
    stroka += 1
    print(fio, cl)
